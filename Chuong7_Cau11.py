import openpyxl

class Employee:
    def __init__(self, emp_id, name, age):
        self.emp_id = emp_id
        self.name = name
        self.age = age

    def __str__(self):
        return f"{self.emp_id} - {self.name} - {self.age}"


class EmployeeManager:
    def __init__(self):
        self.employees = []

    # Đọc Excel theo đúng cột: STT | Mã | Tên | Tuổi
    def load_from_excel(self, filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        self.employees.clear()

        # Bỏ dòng tiêu đề
        for row in ws.iter_rows(min_row=2, values_only=True):
            stt, emp_id, name, age = row
            self.employees.append(Employee(emp_id, name, age))

        print("Đã đọc dữ liệu từ Excel!")

    # Sắp xếp theo tuổi tăng dần
    def sort_by_age(self):
        self.employees.sort(key=lambda e: e.age)

    # Lưu lại Excel sau khi sắp xếp
    def save_to_excel(self, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        # Header
        ws.append(["STT", "Mã", "Tên", "Tuổi"])

        # Ghi dữ liệu + tạo lại STT tự động
        for index, emp in enumerate(self.employees, start=1):
            ws.append([index, emp.emp_id, emp.name, emp.age])

        wb.save(filename)
        print("Đã lưu lại Excel!")

    # In danh sách
    def print_all(self):
        for emp in self.employees:
            print(emp)


# ==================== CHẠY THỬ ========================
def main():
    manager = EmployeeManager()

    # ĐỌC FILE
    manager.load_from_excel("nhanvien.xlsx")

    print("Trước khi sắp xếp:")
    manager.print_all()

    # SẮP XẾP
    manager.sort_by_age()

    print("\nSau khi sắp xếp theo tuổi tăng dần:")
    manager.print_all()

    # GHI LẠI FILE
    manager.save_to_excel("nhanvien_sorted.xlsx")


if __name__ == "__main__":
    main()
